from dataclasses import dataclass, field
from enum import Enum

from openpyxl.drawing.image import Image
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook, load_workbook

from SystemInfo import CURRENT_PATH, BUILD_MODE
from Styling import DEFAULT_ALIGNMENT, DEFAULT_BORDER, DEFAULT_FILL, DEFAULT_FONT, StyleConfig, formatWrite, ColorPalette


class SheetDataEnum(Enum):
    Cell       = 1,
    MergedCell = 2, 
    Row        = 3,
    Column     = 4, 
    Table      = 5,
    Image      = 6
# End


@dataclass
class SheetDataEntry:
    dataType: SheetDataEnum = SheetDataEnum.Cell
    
    # Had to use field because making dataclass with a mutable field makes python whiny
    style: StyleConfig = field( default_factory = StyleConfig ( 
            currentFont = DEFAULT_FONT, 
            currentBorder = DEFAULT_BORDER, 
            currentFill = DEFAULT_FILL, 
            currentAlignment = DEFAULT_ALIGNMENT
        )
    )
    
    cellStart: tuple[int, int]     = (1, 1) # row, column
    cellEnd:   tuple[int, int]     = (1, 1) # row, column
    name:      str                 = "Empty Cell"
    tableData: list[list[str]]     = None
    data:      list[str] | None    = None
    size:      tuple[float, float] = None
    anchor:    str                 = None

    def checkEntryType(self):
        match self.dataType:
            case SheetDataEnum.Cell:
                if (self.cellStart != self.cellEnd) or (len(self.data) > 1) or (self.size != None):
                    print("[warn] Content entered may not be consistent with entry Type { ", self.dataType.name, " }")
                pass
            case SheetDataEnum.MergedCell:
                if (self.cellStart == self.cellEnd) or (len(self.data) > 1) or (self.size != None):
                    print("[warn] Content entered may not be consistent with entry Type { ", self.dataType.name, " }")
                pass
            case SheetDataEnum.Row:
                if (self.cellStart == self.cellEnd) or (len(self.data) == 1) or (self.size != None) or (self.cellStart[1] == self.cellEnd[1]):
                    print("[warn] Content entered may not be consistent with entry Type { ", self.dataType.name, " }")
                pass
            case SheetDataEnum.Column:
                if (self.cellStart == self.cellEnd) or (len(self.data) == 1) or (self.size != None) or (self.cellStart[0] == self.cellEnd[0]) :
                    print("[warn] Content entered may not be consistent with entry Type { ", self.dataType.name, " }")            
                pass
            case SheetDataEnum.Table:
                # todo
                (startRow, startColumn) = self.cellStart
                (endRow, endColumn) = self.cellEnd
                if(((endRow - startRow) < 2) or ((endColumn - startColumn) < 2) or (self.tableData == None)):
                    print("[warn] Content entered may not be consistent with entry Type { ", self.dataType.name, " }") 
                pass
            case SheetDataEnum.Image:
                if((self.size == None) or (self.data == None) or (self.anchor == None)):
                    print("[warn] Content entered may not be consistent with entry Type { ", self.dataType.name, " }") 
                pass

    def toCellStr(self, wb: Workbook) -> tuple[str, str]: # start, end
        start = wb.active.cell(row = self.cellStart[0], column = self.cellStart[1]).coordinate
        end = wb.active.cell(row = self.cellEnd[0], column = self.cellEnd[1]).coordinate
        return (start, end)
        
# don't worry about this, just for notes        
"""     
def sumthing():
    xy = coordinate_from_string('A4') # returns ('A',4)
    col = column_index_from_string(xy[0]) # returns 1
    row = xy[1] 
    wb.active.cell(row = 2, column = 2).coordinate
    pass 
"""

# End



class ExcelSheetData(object):
    entryList: list[SheetDataEntry] = []
    filePath: str = None
    workBook = None

    # Basic functions
    def __init__(self, path = CURRENT_PATH, fileName = "sample.xlsx", loadExisting = False) -> None:
        self.filePath = path + fileName
        
        if loadExisting:
            print("[Loading Excel Sheet] :- ", self.filePath, " -: ... ")
            self.workBook = load_workbook(self.filePath)  
        else:
            print("[Creating Excel Sheet] :- ", self.filePath, " -: ...")
            self.workBook = Workbook()

    def save(self):
        self.workBook.save(self.filePath)
    
    def clear(self): # make current workbook file cleared of existing info
        self.workBook = Workbook()
        self.workBook.save(self.filePath)
    # End 


    # Editing functions
    def editRowSize(self, customRowHeights: list[(int, int)]): # int = row number, int = custom height
        for row in customRowHeights:
            self.workBook.active.row_dimensions[row[0]].height = row[1]

    def editColumnSize(self, customColumnWidths: list[(str, int)] ): # str = character for given column, int = custom width
        for column in customColumnWidths:
            self.workBook.active.column_dimensions[column[0]].width = column[1]
    
    def insertEntry(self, data: SheetDataEntry):
        self.entryList.append(data)

    def updateEntries(self):
        print("[Updating Excel Sheet] ...")
        
        for entry in self.entryList:
            entry.checkEntryType()

            match entry.dataType:
                case SheetDataEnum.Cell:
                    formatWrite(self.workBook, entry.style, (entry.toCellStr(self.workBook))[0], entry.data[0])

                case SheetDataEnum.MergedCell:
                    # Must use int coordinates for merge cell function 
                    self.workBook.active.merge_cells(start_row = entry.cellStart[0], start_column = entry.cellStart[1], end_row = entry.cellEnd[0], end_column = entry.cellEnd[1])
                    formatWrite(self.workBook, entry.style, (entry.toCellStr(self.workBook))[0], entry.data[0])

                case SheetDataEnum.Row: # Finally fixed `elif` case, see comments below for explanation.
                    # Note: Must add 1 to cell difference in order for iteration to be complete (Ex: Starts at 1, Ends at 10 but (10 - 1))
                    if (len(entry.data) >= (entry.cellEnd[1] - entry.cellStart[1] + 1)): # If length of data is greater than row space provided
                        for x in range(entry.cellEnd[1] - entry.cellStart[1] + 1): # Iterate through row length 
                            formatWrite(self.workBook, entry.style, (entry.toCellStr(self.workBook))[0], entry.data[x - 1]) # (x - 1) to convert range to array index
                            temp = list(entry.cellStart) # convert cellStart tuple to list in order to edit
                            temp[1] += 1 # step cellStart forward an iteration
                            entry.cellStart = tuple(temp) # convert back to tuple
                    elif len(entry.data) < (entry.cellEnd[1] - entry.cellStart[1] + 1): # Else if data is smaller than row length given
                        for x in range(entry.cellEnd[1] - entry.cellStart[1] + 1): # Iterate through row length
                            if x >= len(entry.data): # If the current iter, `x`, exceeds the data array length 
                                newCell = (entry.cellStart[0], entry.cellEnd[1] + x) # (row remains the same, column += iter)  
                                newEntry = entry # create copy of entry
                                newEntry.cellEnd = newCell # assign cellEnd to copy
                                # Based on script mode fillin empty cells
                                emptyCellStr = ""
                                if BUILD_MODE == 'Debug':
                                    emptyCellStr = "[Undefined Element]"
                                formatWrite(self.workBook, entry.style, (newEntry.toCellStr(self.workBook))[0], emptyCellStr)
                            else:
                                formatWrite(self.workBook, entry.style, (entry.toCellStr(self.workBook))[0], entry.data[x])
                                
                            temp = list(entry.cellStart) # See above
                            temp[1] += 1
                            entry.cellStart = tuple(temp)
                    elif entry.data == None:
                        pass

                case SheetDataEnum.Column: # See `Row` case for logic explanation
                    if (len(entry.data) >= (entry.cellEnd[0] - entry.cellStart[0] + 1)): 
                        for x in range(entry.cellEnd[0] - entry.cellStart[0] + 1): 
                            formatWrite(self.workBook, entry.style, (entry.toCellStr(self.workBook))[0], entry.data[x - 1])
                            temp = list(entry.cellStart)
                            temp[0] += 1 
                            entry.cellStart = tuple(temp) 
                    elif len(entry.data) < (entry.cellEnd[0] - entry.cellStart[0] + 1): 
                        for x in range(entry.cellEnd[0] - entry.cellStart[0] + 1): 
                            if x >= len(entry.data):
                                newCell = (entry.cellStart[0] + x, entry.cellEnd[1])  
                                newEntry = entry 
                                newEntry.cellEnd = newCell
                                emptyCellStr = ""
                                if BUILD_MODE == 'Debug':
                                    emptyCellStr = "[Undefined Element]"
                                formatWrite(self.workBook, entry.style, (newEntry.toCellStr(self.workBook))[0], emptyCellStr)
                            else:
                                formatWrite(self.workBook, entry.style, (entry.toCellStr(self.workBook))[0], entry.data[x])
                                
                            temp = list(entry.cellStart)
                            temp[0] += 1
                            entry.cellStart = tuple(temp)
                    elif entry.data == None:
                        pass

                case SheetDataEnum.Table:
                    for i, value in enumerate(entry.data, start = entry.cellStart[0]):
                        formatWrite(self.workBook, entry.style, self.workBook.active.cell(row = entry.cellStart[0], column = entry.cellStart[1] + 1).coordinate, value)
                    columnRange = entry.cellEnd[0] - entry.cellStart[0]
                    rowRange = entry.cellEnd[1] - entry.cellStart[1]
                    for i, listValue in enumerate(entry.tableData, start = entry.cellStart[0]):
                        for j, value in enumerate(listValue, start = entry.cellStart[1]):
                            row = entry.cellStart[0] + i
                            column = entry.cellStart[1] + j
                            formatWrite(self.workBook, entry.style, self.workBook.active.cell(row = row, column = column).coordinate, value)

                case SheetDataEnum.Image: # Should work, still uncertain, very much "glue together and hope it works"
                    img = Image(entry.data[0]) # Should be image location
                    img.height = entry.size[0] # insert image height in pixels as float or int (e.g. 305.5) 
                    img.width = entry.size[1] # insert image width in pixels as float or int (e.g. 405.8)
                    img.anchor = entry.anchor # where you want image to be anchored/start from 'A1'
                    self.workBook.active.add_image(img) # adding in the image

            self.entryList.remove(entry)
# End
