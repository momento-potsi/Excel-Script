
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from Styling import DEFAULT_ALIGNMENT, DEFAULT_BORDER, DEFAULT_FILL, DEFAULT_FONT, StyleConfig, formatWrite, ColorPalette
from SheetData import SheetDataEntry, SheetDataEnum, ExcelSheetData

import SystemInfo


class StylingUnitTest(object):
    def colorStyling(self): # Test Passed
        style = StyleConfig (
            currentFont = DEFAULT_FONT, 
            currentBorder = DEFAULT_BORDER, 
            currentFill = DEFAULT_FILL, 
            currentAlignment = DEFAULT_ALIGNMENT
        )

        colorList = [
            ColorPalette.WHITE, 
            ColorPalette.YELLOW_HIGHLIGHT, 
            ColorPalette.BRIGHT_GREEN, 
            ColorPalette.PALE_GREEN, 
            ColorPalette.PALE_BLUE, 
            ColorPalette.GRAY,
            ColorPalette.PALE_YELLOW, 
            ColorPalette.PALE_ORANGE,
            ColorPalette.PALE_RED
        ]
        
        wb = Workbook()
        wb.save(SystemInfo.CURRENT_PATH + "sample.xlsx") # to clear file

        for i in range(len(colorList)):    
            testColor: str = colorList[i].value[0]
            # fills in background color
            style.currentFill = PatternFill (fill_type = "solid", start_color = testColor, end_color = '00FFFFFF')

            wb = load_workbook(SystemInfo.CURRENT_PATH + "sample.xlsx")   
            formatWrite(wb, style, 'A' + str(i + 1), str(colorList[i].name))
            wb.save(SystemInfo.CURRENT_PATH + "sample.xlsx")
            print("[Unit Test]: Styling " + colorList[i].value[0])

        """ Output => {
                [Testing]...
                [Unit Test]: Styling 00FFFFFF
                [Unit Test]: Styling FFFFF200
                [Unit Test]: Styling FF72BF44
                [Unit Test]: Styling FFCCFFCC
                [Unit Test]: Styling FFADC5E7
                [Unit Test]: Styling FF808080
                [Unit Test]: Styling FFFFF9AE
                [Unit Test]: Styling FFF9A870
                [Unit Test]: Styling FFF37B70
                Done!
            }        
        """
# End StylingUnitTest
            

class DataEntryTest(object):

<<<<<<< HEAD
    def entryInstanceTest(self): # Test Passed
        wb = Workbook()
        wb.save(SystemInfo.CURRENT_PATH + "sample.xlsx") # to clear file

        testCases = [
            SheetDataEntry( # Cell Test Case
                dataType = SheetDataEnum.Cell,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (1, 1),
                name = "New Cell",
                data = ["Sample Text"],
                size = None
            ),
            SheetDataEntry( # MergedCell Test Case
                dataType = SheetDataEnum.MergedCell,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (2, 1),
                name = "New Merged Cell",
                data = ["Sample Text For A Longer Cell"],
                size = None
            ),
            SheetDataEntry( # Row Test Case
                dataType = SheetDataEnum.Row,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (1, 10),
                name = "New Row",
                data = ["Sample Text", "hi", "there", "tosin"],
                size = None
            ),
            SheetDataEntry( # Column Test Case
                dataType = SheetDataEnum.Column,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (10, 1),
                name = "New Cell",
                data = ["Sample Text", "hi", "look up", "look up"],
                size = None
            ),
            SheetDataEntry( # Table Test Case
                dataType = SheetDataEnum.Table,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (10, 10),
                name = "New Table",
                data = ["Sample Text", "1+1", "2+2", "2", "4", "..."],
                size = None
            ),
            SheetDataEntry( # Image Test Case
                dataType = SheetDataEnum.Image,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (10, 10),
                name = "New Image",
                data = ["/mnt/D/TemplateImg.png"],
                size = (100, 100),
                anchor = 'A1'
            )
        ]
        
        wb = load_workbook(SystemInfo.CURRENT_PATH + "sample.xlsx")
        
        for i in range(len(testCases)):
            testCases[i].checkEntryType()    
            stringTuple = testCases[i].toCellStr(wb)
            print("[Unit Test] -> Case <" + str(i) + ">: Expected Cell String [" + stringTuple[0] + ", " + stringTuple[1] + "]")
        
        # also test wrong dataType test Cases
        wrongTestCases = [
            SheetDataEntry( # Cell Test Case
                dataType = SheetDataEnum.Table,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (1, 1),
                name = "New Cell",
                data = ["Sample Text"],
                size = None
            ),
            SheetDataEntry( # MergedCell Test Case
                dataType = SheetDataEnum.Row,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (2, 1),
                name = "New Merged Cell",
                data = ["Sample Text For A Longer Cell"],
                size = None
            ),
            SheetDataEntry( # Row Test Case
                dataType = SheetDataEnum.MergedCell,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (1, 10),
                name = "New Row",
                data = ["Sample Text", "hi", "there", "tosin"],
                size = None
            ),
            SheetDataEntry( # Column Test Case
                dataType = SheetDataEnum.Image,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (10, 1),
                name = "New Cell",
                data = ["Sample Text", "hi", "look up", "look up"],
                size = None
            ),
            SheetDataEntry( # Table Test Case
                dataType = SheetDataEnum.Cell,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (10, 10),
                name = "New Table",
                data = ["Sample Text", "1+1", "2+2", "2", "4", "..."],
                size = None
            ),
            SheetDataEntry( # Image Test Case
                dataType = SheetDataEnum.Column,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (10, 10),
                name = "New Image",
                data = ["/mnt/D/TemplateImg.png"],
                size = (100, 100)
            )
        ]

        for i in range(len(wrongTestCases)):
            wrongTestCases[i].checkEntryType()    
            stringTuple = testCases[i].toCellStr(wb)
            print("[Unit Test (checking data types)] -> Wrong Case <" + str(i) + ">: Expected Cell String [" + stringTuple[0] + ", " + stringTuple[1] + "]")

    def excelSheetEditTest(self): # Test Passed
        print("[Unit Test]: Editing Row(s) -> [1 , 3] to have size (10)")
        newSheetStruct = ExcelSheetData()
        (rowNum, height)  = (1, 100)
        newSheetStruct.editRowSize([(rowNum, height)])
        (rowNum, height)  = (3, 100)
        newSheetStruct.editRowSize([(rowNum, height)])

        print("[Unit Test]: Editing Column(s) -> [A , C] to have size (10)")
        (char, width) = ('A', 100)
        newSheetStruct.editColumnSize([(char, width)])        
        (char, width) = ('C', 100)
        newSheetStruct.editColumnSize([(char, width)])

        newSheetStruct.save()
        pass

    def excelSheetUpdateTest(self):
        newSheetStruct = ExcelSheetData()
        INCHES_TO_PIXEL = 0.0104145

        testCases = [
            SheetDataEntry( # Cell Test Case
                dataType = SheetDataEnum.Cell,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (1, 1),
                name = "New Cell",
                data = ["Sample Text"],
                size = None
            ),
            SheetDataEntry( # MergedCell Test Case
                dataType = SheetDataEnum.MergedCell,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (2, 1),
                name = "New Merged Cell",
                data = ["Sample Text For A Longer Cell"],
                size = None
            ),
            SheetDataEntry( # Row Test Case
                dataType = SheetDataEnum.Row,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (1, 10),
                name = "New Row",
                data = ["Sample Text", "hi", "there", "tosin"],
                size = None
            ),
            SheetDataEntry( # Column Test Case
                dataType = SheetDataEnum.Column,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (10, 1),
                name = "New Cell",
                data = ["Sample Text", "hi", "look up", "look up"],
                size = None
            ),
            SheetDataEntry( # Table Test Case
                dataType = SheetDataEnum.Table,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (10, 10),
                name = "New Table",
                data = ["Fruit", "2011", "2012", "2013", "2014"],
                size = None,
                tableData = [
                    ['Apples', str(10000), str(5000), str(8000), str(6000)],
                    ['Pears',   str(2000), str(3000), str(4000), str(5000)],
                    ['Bananas', str(6000), str(6000), str(6500), str(6000)],
                    ['Oranges',  str(500),  str(300),  str(200),  str(700)],
                ]
            ),
            SheetDataEntry( # Image Test Case
                dataType = SheetDataEnum.Image,
                style = StyleConfig (
                    currentFont = DEFAULT_FONT, 
                    currentBorder = DEFAULT_BORDER, 
                    currentFill = DEFAULT_FILL, 
                    currentAlignment = DEFAULT_ALIGNMENT
                ),
                cellStart = (1, 1),
                cellEnd = (10, 10),
                name = "New Image",
                data = ["/mnt/D/TemplateImg.png"],
                size = (2.63 / INCHES_TO_PIXEL, 1.67 / INCHES_TO_PIXEL),
                anchor = 'A1'
            )
        ]

        # todo flesh out table case & finish test
        print("[Unit Test]: Inserting test cases ...")
        for i in range(len(testCases)):
            newSheetStruct.insertEntry(testCases[i])

        print("[Unit Test]: Updating Excel entries ...")
        newSheetStruct.updateEntries()
        
        print("[Unit Test]: Saving Excel Data Sheet")
        newSheetStruct.save()

        pass


=======
    def entryInstanceTest():
        cellType = SheetDataEnum.Cell
        
        pass
>>>>>>> 08692c939fdf1334661611f12797957a44de86f3
# Todo: add commenting, refactor for performance, clean up
