from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from enum import Enum
from dataclasses import dataclass


# Constants
DEFAULT_FONT = Font (
    name = 'Arial',
    size = 12,
    bold = False,
    italic = False,
    vertAlign = None,
    underline = 'none',
    strike = False,
    color = '00000000'
)

DEFAULT_BORDER = Border (
    left = Side(border_style = None, color = 'FF000000'), 
    right = Side(border_style = None, color = 'FF000000'), 
    top = Side(border_style = None, color = 'FF000000'),
    bottom = Side(border_style = None, color = 'FF000000'),
    diagonal = Side(border_style = None, color = 'FF000000'),
    diagonal_direction = 0,
    outline = Side(border_style = None, color = 'FF000000'),
    vertical = Side(border_style = None, color = 'FF000000'),
    horizontal = Side(border_style = None, color = 'FF000000')
)

DEFAULT_FILL = PatternFill (fill_type = None, start_color = 'FFFFFFFF', end_color = 'FF000000')

DEFAULT_ALIGNMENT = Alignment (
    horizontal = 'general',
    vertical = 'bottom',
    text_rotation = 0,
    wrap_text = False,
    shrink_to_fit = False,
    indent = 0
)
# End Constants


# Core Utility
@dataclass
class StyleConfig():
    currentFont: Font           = None
    currentBorder: Border       = None
    currentFill: PatternFill    = None
    currentAlignment: Alignment = None

    def applyConfig(self, workBook: Workbook, cell: str):
        ws = workBook.active
        ws[cell].font = self.currentFont
        ws[cell].border = self.currentBorder
        ws[cell].fill = self.currentFill
        ws[cell].alignment = self.currentAlignment



def formatWrite(workBook: Workbook, style: StyleConfig | None, cell: str, data: str) -> None:
    if style != None:
        style.applyConfig(workBook, cell)
    ws = workBook.active
    if ws[cell].value is not None:
        ws[cell].value += data
    else:
        ws[cell].value = data
# End    


# Template Colors
class ColorPalette(Enum): # Access using `.value`
    WHITE            = '00FFFFFF', # Default color, Note: don't use if just plan text or cell borders will be white
    YELLOW_HIGHLIGHT = 'FFFFF200', # Used on E17 
    BRIGHT_GREEN     = 'FF72BF44', # H17, G30
    PALE_GREEN       = 'FFCCFFCC', # G31
    PALE_BLUE        = 'FFADC5E7', # E18, C45
    GRAY             = 'FF808080', # ROW 20
    PALE_YELLOW      = 'FFFFF9AE', # G32
    PALE_ORANGE      = 'FFF9A870',
    PALE_RED         = 'FFF37B70',


# End Template Colors